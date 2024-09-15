import asyncio
import json
import os
from datetime import datetime
import pytz
from websockets import connect, ConnectionClosed
from termcolor import colored, cprint
from colorama import init
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side

# Initialize Colorama
init()
print("🗄️ Bibliotheken erfolgreich importiert")

# Configuration
symbols = [
    'btcusdt', 'ethusdt', 'solusdt', 'bnbusdt', 'dogeusdt', 'usdcusdt',
    'xrpusdt', 'adausdt', 'maticusdt', 'tonusdt', 'linkusdt', 'trxusdt', 'nearusdt',
    'xlmusdt', 'rndrusdt', 'dotusdt', 'uniusdt', 'atomusdt', 'xmrusdt', 'ldousdt', 'gmxusdt'
]
websocket_url_base_binance = 'wss://fstream.binance.com/ws/'
websocket_url_base_coinbase = 'wss://ws-feed.exchange.coinbase.com'
websocket_url_liq = 'wss://fstream.binance.com/ws/!forceOrder@arr'
xlsx_filename = '/home/nilas/Schreibtisch/Programming/_Algo_Trade_Edge/Backtesting/⭐Big_Trade_and_Liq_History.xlsx'
print("📡 Konfiguration geladen")

# Initialize Maps and Files
name_map = {
    'BTC': '🟡BTC     ', 'ETH': '💠ETH     ', 'SOL': '👾SOL     ', 'BNB': '🔶BNB     ', 'DOGE': '🐶DOGE    ',
    'USDC': '💵USDC    ', 'XRP': '⚫XRP     ', 'ADA': '🔵ADA     ', 'MATIC': '🟣MATIC   ',
    'TON': '🎮TON     ', 'LINK': '🔗LINK    ', 'TRX': '⚙️TRX     ', 'NEAR': '🔍NEAR    ', 'XLM': '🌟XLM     ',
    'RNDR': '🎨RNDR    ', 'DOT': '⚪DOT     ', 'UNI': '🦄UNI     ', 'ATOM': '⚛️ATOM    ', 'XMR': '👽XMR     ',
    'LDO': '🧪LDO     ', 'GMX': '🌀GMX     '
}

cumulative_sum_map = {symbol.upper().replace('USDT', ''): 0 for symbol in symbols}
cumulative_sum_map_liq = {}
print("💾 Maps und Dateien initialisiert")

# Tracking Map for counting trades and liquidations
trade_count_map = {}
connection_closed_logged = set()

# Function to adjust column widths, swap columns B and C, and add borders
def adjust_excel_columns(ws):
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width += 11.32  # Increase by 3 cm

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in range(2, ws.max_row + 1):
        # Swap Symbol and Trade Type
        trade_type = ws.cell(row=row, column=2).value
        symbol = ws.cell(row=row, column=3).value
        ws.cell(row=row, column=2).value = trade_type
        ws.cell(row=row, column=3).value = symbol

        # Color the rows based on Trade Type
        if trade_type == '📈 ':
            fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green fill
        elif trade_type == '📉 ':
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red fill
        else:
            fill = None

        for col in range(1, 8):  # Apply fill and border to columns A to G
            cell = ws.cell(row=row, column=col)
            if fill:
                cell.fill = fill
            cell.border = thin_border

# Function to write data to the Excel file
def write_to_excel(trade_time, symbol, trade_type, price, quantity, usd_size, stars):
    if os.path.exists(xlsx_filename):
        wb = load_workbook(xlsx_filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Trade and Liquidation History"
        ws.append(["Event Time", "Trade Type", "Symbol", "Price", "Quantity", "USD Size", "Stars"])

    symbol_name = name_map.get(symbol.upper().replace('USDT', ''), symbol.upper())
    trade_time_str = datetime.fromtimestamp(trade_time / 1000, pytz.timezone("Europe/Berlin")).strftime('%Y-%m-%d %H:%M:%S')
    ws.append([trade_time_str, trade_type, symbol_name, price, quantity, usd_size, stars])
    adjust_excel_columns(ws)
    wb.save(xlsx_filename)
    wb.close()

# Process Trade Function for Normal Trades
async def process_trade(symbol, price, quantity, trade_time, is_buyer_maker):
    usd_size = price * quantity
    if usd_size <= 5000000:
        return  # Ignore trades below 5,000,000 USD

    SYMBL = symbol.upper().replace('USDT', '')
    trade_type = '📉 ' if is_buyer_maker else '📈 '
    cumulative_sum = update_cumulative_sum(SYMBL, usd_size, trade_type)
    time_berlin = format_trade_time(trade_time)
    output = f"{name_map[SYMBL]}{'|'}{get_stars(usd_size)}{'|'}{time_berlin}{'|'}{trade_type}{format_usd_size(usd_size, trade_type)}{'|'} 💵🟰 {cumulative_sum}"

    if usd_size > 20000000:
        output = add_color_border(output, usd_size)
    elif usd_size < -20000000:
        output = add_color_border(output, usd_size)

    cprint(output, 'white', attrs=get_attrs_trades(usd_size))
    write_to_excel(trade_time, symbol, trade_type, price, quantity, usd_size)
    update_trade_count(SYMBL, get_stars(usd_size), trade_type)

def add_color_border(text, usd_size):
    color = 'green' if usd_size > 0 else 'red'
    lines = text.split('\n')
    width = max(len(line) for line in lines)
    top_border = colored('+' + '-' * (68) + '+', color, attrs=['bold'])
    bottom_border = colored('+' + '-' * (68) + '+', color, attrs=['bold'])
    bordered_text = top_border + '\n'
    for line in lines:
        bordered_text += colored('|', color, attrs=['bold']) + ' ' + line + ' ' * (width - len(line)) + colored('|', color, attrs=['bold']) + '\n'
    bordered_text += bottom_border
    return bordered_text

def get_stars(usd_size):
    if usd_size >= 500000000:
        return '❓💰🃏💰❓'
    elif usd_size >= 300000000:
        return '💸🌈🦄🌈💸'
    elif usd_size >= 150000000:
        return '  🐳🐳🐳  '
    elif usd_size >= 80000000:
        return '   🐳🐳   '
    elif usd_size >= 50000000:
        return '    🐳    '
    elif usd_size >= 30000000:
        return '  🦈🦈🦈  '
    elif usd_size >= 20000000:
        return '   🦈🦈   '
    elif usd_size >= 10000000:
        return '    🦈    '
    elif usd_size >= 5000000:
        return '🐠🐠🐠🐠🐠'
    elif usd_size >= 2500000:
        return ' 🐠🐠🐠🐠 '
    elif usd_size >= 1500000:
        return '  🐠🐠🐠  '
    elif usd_size >= 1000000:
        return '   🐠🐠   '
    else:
        return '    🐠    '

def get_attrs_trades(usd_size):
    if usd_size >= 5000000:
        return ['bold', 'blink']
    elif usd_size >= 1000000:
        return ['bold']
    else:
        return []

def update_cumulative_sum(symbol, usd_size, trade_type):
    if trade_type == '📉 ':
        cumulative_sum_map[symbol] -= usd_size
    else:
        cumulative_sum_map[symbol] += usd_size
    return cumulative_sum_map[symbol]

def format_usd_size(usd_size, trade_type):
    usd_size_color = 'green' if trade_type == '📈 ' else 'red'
    return colored(f"{usd_size:,.0f}$", usd_size_color)

def format_trade_time(trade_time):
    berlin = pytz.timezone("Europe/Berlin")
    return datetime.fromtimestamp(trade_time / 1000, berlin).strftime('%H:%M:%S')

def update_trade_count(formatted_symbol, stars, trade_type):
    key = (formatted_symbol, stars, trade_type)
    if key not in trade_count_map:
        trade_count_map[key] = 0
    trade_count_map[key] += 1

# Process Liquidation Function
async def process_liquidation(symbol, side, timestamp, usd_size):
    if usd_size <= 250000:
        return  # Ignore liquidations below 250,000 USD

    SYMBL = symbol.upper().replace('USDT', '')
    liquidation_type = '📈 ' if side == 'SELL' else '📉 '
    cumulative_sum = update_cumulative_sum_liq(SYMBL, usd_size, liquidation_type)
    output = f"{name_map[SYMBL]}{'|'}{get_liq_stars(usd_size)}{'|'}{format_trade_time(timestamp)}{'|'}{liquidation_type}{format_usd_size(usd_size, liquidation_type)}{'|'} 💧🟰 {cumulative_sum}"

    if usd_size > 5000000:
        output = add_color_border(output, usd_size)
    elif usd_size < -5000000:
        output = add_color_border(output, usd_size)

    cprint(output, 'white', attrs=get_attrs_liquidations(usd_size))
    write_to_excel(timestamp, SYMBL, liquidation_type, None, None, usd_size)
    update_trade_count(SYMBL, get_liq_stars(usd_size), liquidation_type)

def update_cumulative_sum_liq(symbol, usd_size, liquidation_type):
    if liquidation_type == '📉':
        cumulative_sum_map_liq[symbol] -= usd_size
    else:
        cumulative_sum_map_liq[symbol] += usd_size
    return cumulative_sum_map_liq[symbol]

def get_liq_stars(usd_size):
    if usd_size > 50000000:
        return '🌊💰🤿💰🌊'
    elif usd_size > 25000000:
        return '💸🌊🤿🌊💸'
    elif usd_size > 10000000:
        return '  🌊🤿🌊  '
    elif usd_size > 5000000:
        return '    🤿    '
    elif usd_size > 2500000:
        return '  🌊🌊🌊  '
    elif usd_size > 1000000:
        return '   🌊🌊   '
    elif usd_size > 500000:
        return '    🌊    '
    elif usd_size > 250000:
        return '💦💦💦💦💦'
    elif usd_size > 100000:
        return ' 💦💦💦💦 '
    elif usd_size > 50000:
        return '  💦💦💦  '
    elif usd_size > 25000:
        return '   💦💦   '
    else:
        return '    💦    '

def get_attrs_liquidations(usd_size):
    if usd_size >= 250000:
        return ['bold', 'blink']
    elif usd_size >= 100000:
        return ['bold']
    else:
        return []

# Binance Trade Stream
async def binance_trade_stream(uri, symbol):
    while True:
        try:
            async with connect(uri, max_size=None) as websocket:
                while True:
                    message = await websocket.recv()
                    data = json.loads(message)
                    event_time = int(data['E'])
                    price = float(data['p'])
                    quantity = float(data['q'])
                    trade_time = int(data['T'])
                    is_buyer_maker = data['m']
                    await process_trade(symbol, price, quantity, trade_time, is_buyer_maker)
        except ConnectionClosed as e:
            if symbol not in connection_closed_logged:
                print(f"📡 ❗🛰️ Connection closed for {symbol}: {e}.  5 Seconds 🔃")
                connection_closed_logged.add(symbol)
            await asyncio.sleep(5)
        except Exception as e:
            print(f"❗Error❗: {e}.  5 Seconds 🔃")
            await asyncio.sleep(5)

# Coinbase Trade Stream
async def coinbase_trade_stream(uri):
    while True:
        try:
            print(f"📶  Coinbase Trades » {uri}")
            async with connect(uri, max_size=None) as websocket:
                subscribe_message = {
                    "type": "subscribe",
                    "channels": [{"name": "ticker"}]
                }
                await websocket.send(json.dumps(subscribe_message))

                while True:
                    message = await websocket.recv()
                    data = json.loads(message)
                    if data['type'] == 'ticker':
                        product_id = data['product_id']
                        symbol = product_id.split('-')[0]
                        price = float(data['price'])
                        quantity = float(data['last_size'])
                        trade_time = int(datetime.fromisoformat(data['time'].replace('Z', '+00:00')).timestamp() * 1000)
                        is_buyer_maker = data['side'] == 'sell'
                        usd_size = price * quantity
                        if usd_size > 1000000:
                            await process_trade(symbol, price, quantity, trade_time, is_buyer_maker)
        except ConnectionClosed as e:
            print(f"📡 ❗ 🛰️: {e}.  5 Seconds 🔃")
            await asyncio.sleep(5)
        except Exception as e:
            print(f"❗Error❗: {e}.  5 Seconds 🔃")
            await asyncio.sleep(5)

# Binance Liquidation Stream
async def binance_liquidation(uri):
    while True:
        try:
            print(f"📶  Binance Liquidation » {uri}")
            async with connect(uri, max_size=None) as websocket:
                while True:
                    msg = await websocket.recv()
                    order_data = json.loads(msg)['o']
                    symbol = order_data['s'].replace('USDT', '')
                    side = order_data['S']
                    timestamp = int(order_data['T'])
                    filled_quantity = float(order_data['z'])
                    price = float(order_data['p'])
                    usd_size = filled_quantity * price

                    if usd_size > 250000:
                        await process_liquidation(symbol, side, timestamp, usd_size)
        except ConnectionClosed as e:
            print(f"📡 ❗ 🛰️: {e}.  5 Seconds 🔃")
            await asyncio.sleep(5)
        except Exception as e:
            print(f"❗Error❗: {e}.  5 Seconds 🔃")
            await asyncio.sleep(5)

# Main Function
async def main():
    print("🔧Starte Hauptfunktion")
    tasks = []
    for symbol in symbols:
        stream_url = f"{websocket_url_base_binance}{symbol}@aggTrade"
        tasks.append(binance_trade_stream(stream_url, symbol))
    tasks.append(coinbase_trade_stream(websocket_url_base_coinbase))
    tasks.append(binance_liquidation(websocket_url_liq))
    print("🔧Starte asyncio.gather")
    await asyncio.gather(*tasks)

print("🔧Starte asyncio.run(main())")
asyncio.run(main())
print("Programm abgeschlossen")
